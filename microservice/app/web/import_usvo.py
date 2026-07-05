"""Импорт карточек УСВО из Excel, загруженного через веб-кабинет, + шаблон-пример.

Формат таблицы свободный: распознаются любые столбцы. Особый столбец — «История
взаимодействия» (свободный текст) — выносится отдельно и нормализуется в события
(history_ai.normalize_history). Остальные столбцы сохраняются как поля карточки.
"""
from __future__ import annotations

import io

import openpyxl

from app.web.usvo import _fmt, _maybe_excel_date, find_field_value

# Рекомендуемый набор столбцов для шаблона-примера. Импорт НЕ требует именно их —
# распознаётся любой заголовок; это лишь удобная заготовка для оператора.
TEMPLATE_COLUMNS = [
    "ФИО УСВО",
    "Дата рождения",
    "Статус",
    "Телефон",
    "Адрес регистрации",
    "Семейное положение",
    "Сведения о родственниках",
    "Дети",
    "Ветеран боевых действий",
    "Награды",
    "Образование",
    "Специальность / профессия",
    "Нуждается в трудоустройстве",
    "Время героя",
    "Герои Подмосковья",
    "Ассоциация ветеранов",
    "В каких мерах поддержки нуждается",
    "Ответственный от администрации",
    "Дата обзвона",
    "История взаимодействия",
]

# Примеры строк (вторая — со «свободной», неструктурированной историей, чтобы
# показать, как ИИ/нормализатор разложит её на оформленные события).
TEMPLATE_ROWS = [
    [
        "Смирнов Алексей Иванович", "14.07.1987", "ветеран боевых действий",
        "+7 (903) 555-21-43", "Московская область, Ленинский г.о., г. Видное, ул. Школьная, д. 12, кв. 7",
        "женат", "Смирнова Анна Петровна (супруга)", "сын, 2015 г.р.", "да",
        "Медаль «За отвагу»", "среднее профессиональное", "водитель", "не нуждается",
        "участник", "нет", "состоит", "оформление выплат и льгот", "Иванова О. П.",
        "02.06.2026",
        "12.01.2026 поставлен на учёт, заведена карточка. 20.01 обзвон — уточнили "
        "состав семьи. 03.03.2026 прошёл реабилитацию в госпитале, выдали "
        "направление на ТСР. Записан на санаторий на апрель. Трудоустроен "
        "водителем с 20.03.2026. Подал документы на компенсацию ЖКУ — на рассмотрении.",
    ],
    [
        "Кузнецов Дмитрий Сергеевич", "29.11.1992", "мобилизованный",
        "+7 (916) 778-04-19", "Московская область, Ленинский г.о., пос. Развилка, ул. Дачная, д. 4",
        "не женат", "", "нет", "оформляется", "нет сведений", "высшее", "электрик",
        "нуждается", "нет", "нет", "не состоит", "помощь в трудоустройстве",
        "Петров С. А.", "10.05.2026",
        "Обратился за помощью в трудоустройстве. Отказ в выплате — нужна справка о "
        "составе семьи. Приглашён в ассоциацию ветеранов, личный приём запланирован "
        "на 25.06.2026.",
    ],
]

# Столбцы, не входящие в карту полей, — по заголовку определяем «Историю».


def _is_history_header(header: str) -> bool:
    h = (header or "").lower()
    return "истор" in h and "взаимод" in h


def parse_usvo_xlsx(file_bytes: bytes) -> list[dict]:
    """Парсит загруженный Excel в список карточек.

    Каждая карточка: name/status/phone/address/birth_date/call_date/awards,
    `fields` (label→value всех прочих столбцов) и `history_raw` (свободный текст).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows()
    try:
        header_cells = next(rows)
    except StopIteration:
        wb.close()
        return []

    headers = [" ".join(str(c.value or "").split()) or f"Колонка {i+1}"
               for i, c in enumerate(header_cells)]
    history_idx = next((i for i, h in enumerate(headers) if _is_history_header(h)), None)

    cards: list[dict] = []
    for row in rows:
        values = [_fmt(c.value, c.number_format) for c in row]
        # Пустая строка — пропускаем.
        if not any((v or "").strip() for v in values):
            continue
        # Доводим до длины заголовков.
        while len(values) < len(headers):
            values.append("")

        fields: list[dict] = []
        history_raw = ""
        for i, header in enumerate(headers):
            val = values[i] if i < len(values) else ""
            if i == history_idx:
                history_raw = val
                continue
            if "дата" in header.lower():
                val = _maybe_excel_date(val)
            if (val or "").strip():
                fields.append({"label": header, "value": val})

        flist = [(f["label"], f["value"]) for f in fields]
        card = {
            "name": find_field_value(flist, "name"),
            "status": find_field_value(flist, "status"),
            "phone": find_field_value(flist, "phone"),
            "address": find_field_value(flist, "address"),
            "birth_date": find_field_value(flist, "birth"),
            "call_date": find_field_value(flist, "call_date"),
            "awards": find_field_value(flist, "awards"),
            "fields": fields,
            "history_raw": history_raw.strip(),
        }
        # Карточку без имени всё равно принимаем (имя достроится синтетикой),
        # но совсем пустую (только заголовки) — нет.
        if card["fields"] or card["history_raw"]:
            cards.append(card)

    wb.close()
    return cards


def build_template_xlsx() -> bytes:
    """Готовый Excel-шаблон с заголовками и двумя примерами заполнения."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Карточки УСВО"

    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="0D4CD3")
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        # Ширина: «История взаимодействия» — пошире.
        width = 60 if name == "История взаимодействия" else max(16, min(34, len(name) + 4))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    for r, example in enumerate(TEMPLATE_ROWS, start=2):
        for col, val in enumerate(example, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()
