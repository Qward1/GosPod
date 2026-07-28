"""Выгрузка данных кабинета в Excel (openpyxl)."""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="0D4CD3")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _yn(value) -> str:
    return "да" if value else "нет"


def _autosize(ws, max_width: int = 70) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        longest = 0
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            longest = max(longest, max((len(s) for s in str(v).splitlines()), default=0))
        ws.column_dimensions[letter].width = min(max_width, max(12, longest + 2))


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    _autosize(ws)


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _history_summary(rec) -> str:
    events = rec.history if rec.history is not None else []
    parts = []
    for e in events:
        date = e.get("date") or ""
        title = e.get("title") or e.get("detail") or ""
        status = e.get("status") or ""
        parts.append(f"{date} {title} ({status})".strip())
    if parts:
        return "\n".join(parts)
    return rec.history_raw or ""


def usvo_xlsx(records: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Карточки УСВО"
    headers = [
        "ID", "ФИО", "Статус", "Дата рождения", "Телефон", "Адрес регистрации",
        "Дата обзвона", "Награды", "Ветеран БД", "Нуждается в трудоустройстве",
        "Давно без связи", "В организациях", "Меры поддержки", "Источник",
        "История взаимодействия",
    ]
    rows = []
    for r in records:
        flags = r.flags or {}
        org = flags.get("org_vremya") or flags.get("org_geroi_mo") or flags.get("org_assoc")
        rows.append([
            r.id, r.name, r.status, r.birth_date, r.phone, r.address,
            r.call_date, r.awards, _yn(flags.get("vbd")), _yn(flags.get("unemployed")),
            _yn(flags.get("stale_contact")), _yn(org), flags.get("support_need") or "",
            "загружено" if r.source == "uploaded" else "таблица",
            _history_summary(r),
        ])
    _write_sheet(ws, headers, rows)
    return _save(wb)


def appeals_xlsx(appeals: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Обращения"
    headers = [
        "ID", "Дата и время", "Тематика", "Тональность", "Суть (ИИ)",
        "Вопрос гражданина", "Статус", "Ответственный", "Ответ",
        "Гражданин", "MAX ID", "Источник",
    ]
    status_label = {"open": "На рассмотрении", "answered": "Отвечено"}
    rows = []
    for a in appeals:
        cit = a.get("citizen") or {}
        senti = a.get("sentiment") or {}
        rows.append([
            a.get("id"), a.get("created_human"), a.get("topic"),
            senti.get("label") or "", a.get("summary") or "",
            a.get("question"), status_label.get(a.get("status"), a.get("status")),
            a.get("assignee") or "", a.get("answer") or "",
            cit.get("name") or "", cit.get("user_id") or "",
            "MAX" if a.get("real") else "демо-данные",
        ])
    _write_sheet(ws, headers, rows)
    return _save(wb)


def applications_xlsx(apps: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявления"
    headers = [
        "ID", "Мера поддержки", "Заявитель", "Категория", "Код",
        "Статус", "Поступило", "Недостающие документы", "Способ выплаты", "Решение принял",
    ]
    rows = []
    for a in apps:
        ap = a.get("applicant") or {}
        pay = a.get("payment") or {}
        pay_str = ("почтой" if pay.get("method") == "post"
                   else f"{pay.get('bank') or ''} {pay.get('account') or ''}".strip())
        rows.append([
            a.get("id"), a.get("measure_title"),
            ap.get("fio") or (a.get("citizen") or {}).get("name") or "",
            a.get("category"), a.get("category_code"),
            a.get("status_label") or a.get("status"), a.get("created_human"),
            ", ".join(a.get("missing") or []), pay_str, a.get("decided_by") or "",
        ])
    _write_sheet(ws, headers, rows)
    return _save(wb)


def analytics_xlsx(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    # Лист 1 — ключевые показатели.
    ws = wb.active
    ws.title = "Показатели"
    appeals = data.get("appeals") or {}
    orgs = data.get("orgs") or {}
    metrics = [
        ["Обращений за день", appeals.get("day", 0)],
        ["Обращений за неделю", appeals.get("week", 0)],
        ["Обращений за месяц", appeals.get("month", 0)],
        ["Очных обращений в администрацию", data.get("in_person", 0)],
        ["Заявлений на меры поддержки", data.get("applications", 0)],
        ["Всего УСВО на учёте", data.get("total_usvo", 0)],
        ["Нуждаются в трудоустройстве", data.get("unemployed", 0)],
        ["Без контакта дольше нормы", data.get("stale_contacts", 0)],
        ["Охват ветеранскими организациями, %", orgs.get("coverage_pct", 0)],
    ]
    _write_sheet(ws, ["Показатель", "Значение"], metrics)

    ws2 = wb.create_sheet("Тематики обращений")
    _write_sheet(ws2, ["Тематика", "Количество"],
                 [[t["label"], t["count"]] for t in data.get("topics", [])])

    ws3 = wb.create_sheet("Меры поддержки")
    _write_sheet(ws3, ["Мера поддержки", "Количество"],
                 [[t["label"], t["count"]] for t in data.get("support_measures", [])])

    hm = data.get("heatmap") or {}
    ws4 = wb.create_sheet("Очаги обращений")
    _write_sheet(ws4, ["Населённый пункт", "Обращений", "Плотность, %"],
                 [[h["name"], h["count"], round(h.get("intensity", 0) * 100)]
                  for h in hm.get("hotspots", [])])
    return _save(wb)
